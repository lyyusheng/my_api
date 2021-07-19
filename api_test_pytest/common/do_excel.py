from openpyxl import load_workbook
from common import project_path
from common.read_config import ReadConfig

"""此模块用于从Excel中读取测试用例"""


class DoExcel:

    def __init__(self, file_name, sheet_name):
        """file_name:要读取的excel文件名
            sheet_name:excel表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_case(self, section):
        """此方法用于读取excel中的用例
            section:配置文件case.conf中的section,用于控制读取哪一部分用例
        """
        # 配置文件控制读取哪些用例 # get_data()是自己写的read_config模块中用于读取配置文件中元组、字典、列表的函数
        case_id = ReadConfig(project_path.conf_path).get_data(section,
                                                              'case_id')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        # sheet = wb(self.sheet_name)   #错
        # sheet = wb.sheetnames(self.sheet_name)  # 错了
        # 空列表存储整个表单的测试用以
        test_data = []
        for i in range(2, sheet.max_row + 1):
            # 空字典存储一条测试用例
            row_data = {}
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Method'] = sheet.cell(i, 4).value
            row_data['Url'] = sheet.cell(i, 5).value
            if sheet.cell(i, 6).value.find("tel") != -1:
                tel = self.get_tel()
                row_data['Params'] = sheet.cell(i, 6).value.replace("tel", str(tel))
                self.updata_tel(int(tel) + 1)

            else:
                row_data['Params'] = sheet.cell(i, 6).value
            row_data['sql'] = sheet.cell(i, 7).value
            row_data['ExpectedResult'] = sheet.cell(i, 8).value
            row_data['ActualResult'] = sheet.cell(i, 9).value
            # test_data = test_data.append(row_data)
            test_data.append(row_data)
        wb.close()
        # 根据配置文件读取测试用例
        final_data = []  # 定义一个空列表，用于存储根据配置文件需要读取的测试用例
        if case_id == "all":  # 把1换成all时报错：不可迭代,配置文件把all变成"all"即可，本来是str,都出来也是str，相当于双重str
            final_data = test_data
        else:
            for i in case_id:  # case_id是配置文件[CASE]里面的section,具体值可自行配置
                final_data.append(test_data[i - 1])  # i是case，从1开始，但是列表索引是从0开始遍历的，所以要减1,但是如果配置文件中的i是"all",
        return final_data

    def write_back(self, row, col, value):
        """写会测试数据"""
        wb = load_workbook(self.file_name)
        # sheet = wb.sheetnames(self.sheet_name)   # 错了
        # sheet = wb(self.sheet_name)   # 也错了，不是（），是()
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = value
        wb.save(self.file_name)
        wb.close()

    def get_tel(self):
        """测函数用于获取excel上的手机号"""
        wb = load_workbook(self.file_name)
        sheet = wb["tel"]
        tel = sheet.cell(1, 2).value
        wb.close()
        return tel

    def updata_tel(self, new_tel):
        """此函数用于自动更新excel上的手机号"""
        wb = load_workbook(self.file_name)
        sheet = wb["tel"]
        sheet.cell(1, 2, new_tel)
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    do = DoExcel(project_path.case_path, "register")
    res = do.read_case('RegisterCase')
    print(res)
